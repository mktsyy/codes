#-*-coding:utf-8-*-
from openpyxl import load_workbook
import when

##装载表格
wb = load_workbook(filename = '58发房.xlsx')

# sheet_ranges = wb['6月22日']

##打印worksheet
# print (wb.get_sheet_names())

##确定当日时间
i = when.today()
today = str(i.month) + "月" + str(i.day) + "日"
# print (today)

##如果当日sheet未建立的话，就建立
if today not in wb.get_sheet_names():
	ws2 = wb.create_sheet(title="6月23日")

##选择今日sheet
sheet_ranges = wb[today]
# print (sheet_ranges)

##打印I21值
# print (type(sheet_ranges['I21'].value))

##赋值表格J21
# sheet_ranges['J21'].value = 99
# print(sheet_ranges['J21'].value)

##保存表格
# wb.save(filename = '58发房.xlsx')

excelNum = [3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18]
region = [
"徐汇",
"长宁",
"浦东",
"浦东",
"杨浦",
"普陀",
"闵行",
"闵行",
"闸北",
"宝山",
"宝山",
"嘉定",
"嘉定",
"松江",
"松江",
"青浦",
]
column = dict(zip(region,excelNum))
print (column)

values = 0
for grip in ["E","F","G","H","I","J","K","L","M","N","O","P"]:
	if isinstance(sheet_ranges[str(grip)+str(column['长宁'])].value,int):
		values += sheet_ranges[str(grip)+str(column['长宁'])].value
print (values)
